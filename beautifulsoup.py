from bs4 import BeautifulSoup
import re
from pprint import pprint

s1_structure = []
s2_structure = []

S1_TAG = "sect1 data-line-*"
S2_TAG = "sect2 data-line-*"
IMPORTANT_TAG = "admonitionblock important *"
CHECKLIST_TAG = "ulist checklist data-line-*"



with open('手順書モデル.html', 'r', encoding='utf-8') as html:
    soup = BeautifulSoup(html, features='html.parser')
    for s1 in soup.find_all(class_=re.compile(S1_TAG)):

        if s1.find(class_=re.compile(IMPORTANT_TAG)) is None:
            important_lines = []
        else:
            important_lines = [ i.strip() for i in s1.find(class_=re.compile(IMPORTANT_TAG)).get_text().splitlines() ]
        s1_structure.append({s1.find('h2').get_text() : {"IMPORTANT" : "\n".join([ l for l in important_lines if l ])}})

        if s1.find(class_=re.compile(S2_TAG)) is None:
            pass
        else:
            s2_structure = []
            for s2 in s1.find_all(class_=re.compile(S2_TAG)):
                if s2.find(class_=re.compile(IMPORTANT_TAG)) is None:
                    important_lines = []
                else:
                    important_lines = [ i.strip() for i in s2.find(class_=re.compile(IMPORTANT_TAG)).get_text().splitlines() ]

                if s2.find(class_=re.compile(CHECKLIST_TAG)) is None:
                    checklist_lines = []
                else:
                    checklist_lines = [ i.strip() for i in s2.find(class_=re.compile(CHECKLIST_TAG)).get_text().splitlines() ]

                operation_lines = [ i.strip() for i in s2.get_text().splitlines() ]
                s2_structure.append({s2.find('h3').get_text() : {"IMPORTANT" : "\n".join([ l for l in important_lines if l ]),"OPERATION" : "\n".join([ l for l in operation_lines if l]),"CHECKLIST" : "\n".join([ l for l in checklist_lines if l])}})
        s1_structure[-1][s1.find('h2').get_text()]["OPERATIONS"] = s2_structure

pprint(s1_structure)
pprint(s2_structure)

from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill

def write_operations_to_cell(excelfilename, sheettitle, cell_num, operation, weapTextopt=True, fill_line_color=None):
    wb = load_workbook(excelfilename)
    ws = wb.active
    ws.title = sheettitle
    ws[cell_num].alignment = Alignment(wrapText=weapTextopt, vertical='center')
    ws[cell_num] = operation
    if fill_line_color:
        for i in range(12):
            color_fill_cellnum = "".join([chr(65+i), cell_num[1:]])
            ws[color_fill_cellnum].fill = PatternFill(fill_type='solid',
                                fgColor=fill_line_color)
    wb.save(filename = excelfilename)


FIRST_CELL_ROWPOSITION = 6
cell_position = FIRST_CELL_ROWPOSITION

operation_pattern_dict = {"本手順は、運用監視端末から実行する": "運用監視端末", "本手順は、HCサーバ1号機から実行する": "HCサーバ1号機","本手順は、アクセス管理サーバから実行する": "ACS"}

for s1 in s1_structure:
    for k,v in s1.items():
        write_operations_to_cell('operation_man_format.xlsx',"詳細手順","".join(["A",str(cell_position)]), k, weapTextopt=False, fill_line_color="FFFF00")
        cell_position += 1
        write_operations_to_cell('operation_man_format.xlsx',"詳細手順","".join(["A",str(cell_position)]), v["IMPORTANT"], weapTextopt=False, fill_line_color="FFFFCC")
        cell_position += 1
        write_operations_to_cell('operation_man_format.xlsx',"詳細手順","".join(["A",str(cell_position)]), k, weapTextopt=False, fill_line_color="CCFFCC")
        cell_position += 1
        for s2_operation in v["OPERATIONS"]:
            write_operations_to_cell('operation_man_format.xlsx',"詳細手順","".join(["B",str(cell_position)]), list(s2_operation.keys())[0])
            write_operations_to_cell('operation_man_format.xlsx',"詳細手順","".join(["C",str(cell_position)]), s2_operation[list(s2_operation.keys())[0]]["IMPORTANT"])
            write_operations_to_cell('operation_man_format.xlsx',"詳細手順","".join(["D",str(cell_position)]), "□")
            for operation_pattern,operation_srv_name in operation_pattern_dict.items():
                if operation_pattern in s2_operation[list(s2_operation.keys())[0]]["IMPORTANT"]:
                    write_operations_to_cell('operation_man_format.xlsx',"詳細手順","".join(["F",str(cell_position)]), operation_srv_name)
                    break
            write_operations_to_cell('operation_man_format.xlsx',"詳細手順","".join(["G",str(cell_position)]), s2_operation[list(s2_operation.keys())[0]]["OPERATION"])
            write_operations_to_cell('operation_man_format.xlsx',"詳細手順","".join(["H",str(cell_position)]), " ".join(["□", s2_operation[list(s2_operation.keys())[0]]["CHECKLIST"]]))
            write_operations_to_cell('operation_man_format.xlsx',"詳細手順","".join(["I",str(cell_position)]), "□")
            cell_position += 1
